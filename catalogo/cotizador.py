from __future__ import annotations

import csv
import os
from dataclasses import dataclass, asdict
from typing import Iterable, List, Dict, Any

try:
	import pandas as pd  # type: ignore
except Exception:  # pragma: no cover - entorno sin pandas
	pd = None  # type: ignore

from .data import cargar_catalogo_simulado, Producto


@dataclass
class ItemCotizacion:
	"""Elemento de la cotización generada."""
	id_convenio_marco: str
	nombre_producto: str
	precio_unitario: float
	cantidad: int
	url_imagen: str
	subtotal: float


def _leer_archivo_generico(ruta_archivo: str) -> List[Dict[str, Any]]:
	"""Lee un archivo CSV, XLSX o TXT y retorna una lista de dicts con claves
	`entrada` y `cantidad`.

	- Para CSV/XLSX: se asume que la primera columna es ID o Nombre.
	- Opcionalmente, si existe una segunda columna numérica, se usa como cantidad.
	- Para TXT: una entrada por línea, admite formato `valor` o `valor,cantidad`.
	"""
	_, ext = os.path.splitext(ruta_archivo.lower())
	if ext == ".csv":
		# Intentar con pandas si está disponible
		if pd is not None:
			df = pd.read_csv(ruta_archivo)
			return _normalizar_df_a_entradas(df)
		# Fallback con csv
		entradas: List[Dict[str, Any]] = []
		with open(ruta_archivo, newline="", encoding="utf-8") as f:
			reader = csv.reader(f)
			filas = list(reader)
			if not filas:
				return []
			# Asumir encabezados en primera fila
			for fila in filas[1:]:
				if not fila:
					continue
				entrada = str(fila[0]).strip()
				cantidad = 1
				if len(fila) > 1:
					try:
						cantidad = int(fila[1])
					except Exception:
						cantidad = 1
				entradas.append({"entrada": entrada, "cantidad": cantidad})
		return entradas
	elif ext in (".xlsx", ".xls"):
		if pd is not None:
			df = pd.read_excel(ruta_archivo)
			return _normalizar_df_a_entradas(df)
		# Fallback con openpyxl
		from openpyxl import load_workbook  # type: ignore

		wb = load_workbook(ruta_archivo)
		ws = wb.active
		entradas: List[Dict[str, Any]] = []
		primera = True
		for row in ws.iter_rows(values_only=True):
			if primera:
				primera = False
				continue
			if not row:
				continue
			entrada = str(row[0]).strip() if row[0] is not None else ""
			if not entrada:
				continue
			cantidad = 1
			if len(row) > 1 and row[1] is not None:
				try:
					cantidad = int(row[1])
				except Exception:
					cantidad = 1
			entradas.append({"entrada": entrada, "cantidad": cantidad})
		return entradas
	elif ext == ".txt":
		entradas: List[Dict[str, Any]] = []
		with open(ruta_archivo, "r", encoding="utf-8") as f:
			for linea in f:
				linea = linea.strip()
				if not linea:
					continue
				# Permite "valor" o "valor,cantidad"
				partes = [p.strip() for p in linea.split(",")]
				entrada = partes[0]
				cantidad = int(partes[1]) if len(partes) > 1 and partes[1].isdigit() else 1
				entradas.append({"entrada": entrada, "cantidad": cantidad})
		return entradas
	else:
		raise ValueError(f"Formato de archivo no soportado: {ext}")


def _normalizar_df_a_entradas(df) -> List[Dict[str, Any]]:
	# Normaliza CSV/XLSX (pandas) a lista de dicts uniformes
	if df is None or getattr(df, "empty", False):
		return []
	primera_columna = df.columns[0]
	cantidad_col = df.columns[1] if len(df.columns) > 1 else None
	entradas: List[Dict[str, Any]] = []
	for _, row in df.iterrows():
		entrada_valor = str(row[primera_columna]).strip()
		if not entrada_valor:
			continue
		cantidad_valor = 1
		if cantidad_col is not None:
			try:
				cantidad_valor = int(row[cantidad_col])
			except Exception:
				cantidad_valor = 1
		entradas.append({"entrada": entrada_valor, "cantidad": cantidad_valor})
	return entradas


def _buscar_producto(catalogo: Dict[str, Producto], entrada: str) -> Producto | None:
	"""Busca por ID exacto o por nombre contiene (case-insensitive)."""
	entrada_norm = entrada.strip()
	if entrada_norm in catalogo:
		return catalogo[entrada_norm]
	# Búsqueda por nombre contiene
	entrada_lower = entrada_norm.lower()
	for _id, prod in catalogo.items():
		if prod.nombre.lower().find(entrada_lower) != -1:
			return prod
	return None


def procesar_archivo_cotizacion(ruta_archivo: str):
	"""Procesa un archivo del usuario y genera una cotización.

	Retorna un DataFrame con columnas:
	- ID_Convenio_Marco
	- Nombre_Producto
	- Precio_Unitario
	- Cantidad
	- URL_Imagen
	- Subtotal

	Lanza ValueError para formatos no soportados.
	"""
	entradas = _leer_archivo_generico(ruta_archivo)
	catalogo = cargar_catalogo_simulado()

	items: List[ItemCotizacion] = []
	errores: List[str] = []
	for e in entradas:
		entrada = str(e["entrada"]).strip()
		cantidad = int(e.get("cantidad", 1))
		producto = _buscar_producto(catalogo, entrada)
		if producto is None:
			errores.append(entrada)
			continue
		subtotal = producto.precio * cantidad
		items.append(
			ItemCotizacion(
				id_convenio_marco=producto.id_convenio,
				nombre_producto=producto.nombre,
				precio_unitario=producto.precio,
				cantidad=cantidad,
				url_imagen=producto.url_imagen,
				subtotal=subtotal,
			)
		)

	registros = [asdict(i) for i in items]
	# Renombrar claves a formato solicitado
	registros = [
		{
			"ID_Convenio_Marco": r["id_convenio_marco"],
			"Nombre_Producto": r["nombre_producto"],
			"Precio_Unitario": r["precio_unitario"],
			"Cantidad": r["cantidad"],
			"URL_Imagen": r["url_imagen"],
			"Subtotal": r["subtotal"],
		}
		for r in registros
	]

	if pd is not None:
		df_cotizacion = pd.DataFrame(registros)
		# Adjuntar información de errores como atributo para la UI
		df_cotizacion.attrs["no_encontrados"] = errores
		return df_cotizacion
	# Fallback: retornar lista de dicts, y exponer errores junto con ella
	return {"cotizacion": registros, "no_encontrados": errores}


def _cli():
	import argparse
	parser = argparse.ArgumentParser(description="Procesa archivos de cotización y genera una salida tabular.")
	parser.add_argument("ruta_archivo", help="Ruta al archivo de entrada (.csv, .txt, .xlsx)")
	parser.add_argument("--salida", help="Ruta a CSV de salida", default="cotizacion_salida.csv")
	args = parser.parse_args()

	resultado = procesar_archivo_cotizacion(args.ruta_archivo)
	if pd is not None and hasattr(resultado, "to_csv"):
		resultado.to_csv(args.salida, index=False)
		no_encontrados = getattr(resultado, "attrs", {}).get("no_encontrados", [])
	else:
		# resultado es dict con claves cotizacion / no_encontrados
		registros = resultado["cotizacion"]
		no_encontrados = resultado.get("no_encontrados", [])
		# Escribir CSV
		fieldnames = [
			"ID_Convenio_Marco",
			"Nombre_Producto",
			"Precio_Unitario",
			"Cantidad",
			"URL_Imagen",
			"Subtotal",
		]
		with open(args.salida, "w", newline="", encoding="utf-8") as f:
			writer = csv.DictWriter(f, fieldnames=fieldnames)
			writer.writeheader()
			for row in registros:
				writer.writerow(row)
	if no_encontrados:
		print("Advertencia: No se encontraron los siguientes productos:")
		for v in no_encontrados:
			print(" -", v)
	print(f"Cotización guardada en {args.salida}")


if __name__ == "__main__":
	_cli()

